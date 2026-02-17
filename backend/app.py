"""
Telecom Customer Complaint Handling System - Backend
=====================================================
Full backend with auth, chat, tickets, and the original AI chatbot integrated.
"""

import os
import json
import time
import random
import string
from datetime import datetime, timezone, timedelta

from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required, get_jwt_identity
)
from flask_mail import Mail, Message
from openai import AzureOpenAI
from dotenv import load_dotenv

from models import db, bcrypt, User, ChatSession, ChatMessage, Ticket, Feedback, SystemSetting

load_dotenv()

# ─── App Setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", os.urandom(24).hex())
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5432/telecom_complaints"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["JWT_SECRET_KEY"] = os.environ.get("JWT_SECRET", "super-secret-jwt-key-change-in-prod")
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=24)

# ─── Flask-Mail Configuration ─────────────────────────────────────────────
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "True").lower() in ("true", "1", "yes")
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_DEFAULT_SENDER", os.environ.get("MAIL_USERNAME"))

CORS(app, supports_credentials=True)
db.init_app(app)
bcrypt.init_app(app)
jwt = JWTManager(app)
mail = Mail(app)


# ─── Azure OpenAI Configuration ──────────────────────────────────────────────
client = AzureOpenAI(
    api_key="",
    api_version="2023-07-01-preview",
    azure_endpoint="https://entgptaiuat.openai.azure.com"
)
DEPLOYMENT_NAME = os.environ.get("AZURE_DEPLOYMENT_NAME", "gpt-4o-mini")


# ═══════════════════════════════════════════════════════════════════════════════
#  CHATBOT CODE 
# ═══════════════════════════════════════════════════════════════════════════════

TELECOM_MENU = {
    "1": {
        "name": "Mobile Services (Prepaid / Postpaid)",
        "icon": "",
        "description": "Covers all issues related to mobile phone services including voice calls, SMS, mobile data, SIM cards, prepaid recharges, postpaid billing, roaming, number portability, and mobile network coverage.",
        "subprocesses": {
            "1": {"name": "Billing & Payment Issues", "semantic_scope": "Unexpected charges, wrong bill amount, double billing, payment failed but money deducted, recharge not credited, balance deducted without usage, auto-renewal charged, EMI issues on phone, refund not received for telecom services, incorrect tax on bill, bill dispute"},
            "2": {"name": "Network / Signal Problems", "semantic_scope": "No signal, weak signal, call drops, poor network coverage, network congestion, unable to make/receive calls, tower issue, dead zone, indoor coverage problem, 4G/5G not available, network outage in area"},
            "3": {"name": "SIM Card & Activation", "semantic_scope": "New SIM not activated, SIM blocked, SIM damaged, SIM swap, eSIM activation, lost SIM replacement, SIM not detected, PUK locked, KYC verification pending, Aadhaar linking with SIM, SIM upgrade to 4G/5G"},
            "4": {"name": "Data Plan & Recharge Issues", "semantic_scope": "Data not working after recharge, wrong plan activated, data exhausted too quickly, unable to recharge, recharge failed but amount debited, validity not extended, data speed throttled, unlimited plan not giving unlimited data, add-on pack issues, coupon/promo code not working"},
            "5": {"name": "International Roaming", "semantic_scope": "Roaming not working abroad, high roaming charges, incoming calls charged during roaming, data roaming activation, roaming pack not applied, unable to call from foreign country, roaming bill shock, ISD/STD calling issues"},
            "6": {"name": "Mobile Number Portability (MNP)", "semantic_scope": "Want to switch operator, MNP request rejected, porting delay, UPC code not received, number lost during porting, services disrupted after porting, porting to another network, port-out issues"},
            "7": {"name": "Call / SMS Failures", "semantic_scope": "Unable to make calls, calls not connecting, one-way audio, SMS not being delivered, SMS not received, OTP not coming, call going to voicemail, DND (Do Not Disturb) issues, spam calls, call forwarding not working, conference call issues"},
            "8": {"name": "Others", "semantic_scope": ""},
        },
    },
    "2": {
        "name": "Broadband / Internet Services",
        "icon": "",
        "description": "Covers all issues related to wired/wireless broadband, fiber internet, DSL connections, WiFi, and home/office internet services.",
        "subprocesses": {
            "1": {"name": "Slow Speed / No Connectivity", "semantic_scope": "Internet too slow, speed not matching plan, buffering while streaming, downloads very slow, no internet connection, WiFi connected but no internet, speed drops at night, latency/ping too high, speed test showing low results, bandwidth issue"},
            "2": {"name": "Frequent Disconnections", "semantic_scope": "Internet keeps disconnecting, connection drops every few minutes, unstable connection, intermittent connectivity, WiFi drops frequently, connection resets, have to restart router repeatedly, disconnects during video calls"},
            "3": {"name": "Billing & Plan Issues", "semantic_scope": "Wrong broadband bill, overcharged, plan upgrade/downgrade issues, FUP limit reached, auto-debit failed, payment not reflected, want to change plan, hidden charges, installation charges disputed, security deposit refund"},
            "4": {"name": "New Connection / Installation", "semantic_scope": "New broadband connection request, installation delayed, technician not showing up, fiber cable not laid, connection pending, availability check, shift connection to new address, relocation of broadband"},
            "5": {"name": "Router / Equipment Problems", "semantic_scope": "Router not working, WiFi router faulty, modem blinking red, ONT device issue, router overheating, need router replacement, firmware update problem, WiFi range too short, LAN port not working, equipment return"},
            "6": {"name": "IP Address / DNS Issues", "semantic_scope": "Cannot access certain websites, DNS resolution failure, need static IP, IP blocked, website loading error, proxy issues, VPN not working over broadband, port forwarding needed"},
            "7": {"name": "Others", "semantic_scope": ""},
        },
    },
    "3": {
        "name": "DTH / Cable TV Services",
        "icon": "",
        "description": "Covers all issues related to Direct-To-Home television, cable TV, set-top boxes, and TV channel subscriptions.",
        "subprocesses": {
            "1": {"name": "Channel Not Working / Missing", "semantic_scope": "Channel not showing, channel removed from pack, channel black screen, paid channel not available, regional channel missing, HD channel not working, channel list changed, favorite channel gone"},
            "2": {"name": "Set-Top Box Issues", "semantic_scope": "Set-top box not turning on, remote not working, set-top box hanging/freezing, recording not working, set-top box overheating, display error on box, need set-top box replacement, software update stuck, box showing boot loop"},
            "3": {"name": "Billing & Subscription", "semantic_scope": "Wrong DTH bill, subscription expired, auto-renewal issue, pack change charges, NCF charges too high, channel added without consent, refund not received, wallet recharge failed, monthly charges incorrect"},
            "4": {"name": "Signal / Picture Quality", "semantic_scope": "No signal on TV, picture breaking/pixelating, rain causing signal loss, dish alignment needed, weak signal, audio out of sync, color distortion, signal loss at certain times, frozen picture, horizontal lines on TV"},
            "5": {"name": "Package / Plan Changes", "semantic_scope": "Want to change channel pack, upgrade to HD, add premium channels, downgrade plan, customize channel selection, regional pack addition, sports pack subscription, plan comparison, best value pack"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
    "4": {
        "name": "Landline / Fixed Line Services",
        "icon": "",
        "description": "Covers all issues related to traditional landline phone services, fixed-line connections, and wired telephone services.",
        "subprocesses": {
            "1": {"name": "No Dial Tone / Dead Line", "semantic_scope": "Landline not working, no dial tone, line dead, phone silent, no sound when picking up receiver, line suddenly stopped working, connection cut off, cable damaged"},
            "2": {"name": "Call Quality Issues (Noise / Echo)", "semantic_scope": "Static noise on landline, echo during calls, crackling sound, voice breaking, cross-connection hearing other conversations, humming noise, low volume on calls, distorted audio"},
            "3": {"name": "Billing & Charges", "semantic_scope": "Landline bill too high, calls charged incorrectly, wrong number dialed charges, rental overcharged, payment not updated, metered vs unlimited plan dispute, ISD charges on landline"},
            "4": {"name": "New Connection / Disconnection", "semantic_scope": "Want new landline connection, disconnection request, temporary suspension, connection shifting to new address, reconnection after disconnection, transfer of ownership"},
            "5": {"name": "Fault Repair Request", "semantic_scope": "Cable cut in area, junction box damaged, overhead wire fallen, underground cable fault, technician visit needed, repeated fault in same line, wet cable causing issues, maintenance request"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
    "5": {
        "name": "Enterprise / Business Solutions",
        "icon": "",
        "description": "Covers all issues related to business/corporate telecom solutions including leased lines, SLA-based services, bulk connections, cloud telephony, and managed network services.",
        "subprocesses": {
            "1": {"name": "SLA Breach / Service Downtime", "semantic_scope": "Service level agreement not met, uptime guarantee violated, business internet down, prolonged outage affecting business, compensation for downtime, SLA penalty claim, response time exceeded"},
            "2": {"name": "Leased Line / Dedicated Connection", "semantic_scope": "Leased line down, dedicated bandwidth not delivered, point-to-point link failure, MPLS circuit issue, last mile connectivity problem, fiber cut affecting leased line, jitter/latency on dedicated line"},
            "3": {"name": "Bulk / Corporate Plan Issues", "semantic_scope": "Corporate plan benefits not applied, bulk SIM management, employee connection issues, CUG (Closed User Group) problem, corporate billing discrepancy, group plan changes"},
            "4": {"name": "Cloud / VPN / MPLS Issues", "semantic_scope": "VPN tunnel down, MPLS network unreachable, cloud connectivity slow, SD-WAN issue, site-to-site VPN failure, enterprise cloud access problem, managed WiFi for office not working"},
            "5": {"name": "Technical Support Escalation", "semantic_scope": "Need senior technician, previous complaint not resolved, multiple complaints on same issue, want to escalate to manager, technical team not responding, critical issue needs immediate attention"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
}


def get_subprocess_details(sector_key: str) -> str:
    sector = TELECOM_MENU[sector_key]
    details = []
    for k, v in sector["subprocesses"].items():
        if isinstance(v, dict) and v["name"] != "Others":
            details.append(f'SUBPROCESS: "{v["name"]}"\n  Typical issues: {v["semantic_scope"]}')
    return "\n\n".join(details)


def get_subprocess_name(sector_key: str, subprocess_key: str) -> str:
    sector = TELECOM_MENU.get(sector_key, {})
    sp = sector.get("subprocesses", {}).get(subprocess_key, {})
    if isinstance(sp, dict):
        return sp.get("name", "Others")
    return sp if isinstance(sp, str) else "Others"


def is_telecom_related(query: str, sector_name=None, subprocess_name=None) -> bool:
    context_block = ""
    if sector_name:
        context_block = (
            f'\n\n── USER\'S MENU NAVIGATION ──\n'
            f'The user already selected telecom sector: "{sector_name}"'
        )
        if subprocess_name:
            context_block += f'\nThey also selected subprocess: "{subprocess_name}"'
        context_block += (
            "\n\nBecause the user navigated a TELECOM complaint menu to reach this point, "
            "their query is almost certainly telecom-related. Generic complaints like "
            "'money deducted', 'service not working', 'bad experience', 'want refund', "
            "'not getting what I paid for' etc. should be interpreted in the telecom context.\n"
            "Only classify as NOT telecom if the query is EXPLICITLY about a completely "
            "different industry."
        )
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    "You are a semantic intent classifier for a TELECOM complaint chatbot.\n\n"
                    "Your job is to determine whether the user's query is related to telecommunications.\n\n"
                    "TELECOM includes (but is not limited to):\n"
                    "- Mobile phone services (calls, SMS, data, prepaid, postpaid)\n"
                    "- Internet/broadband/WiFi/fiber services\n"
                    "- DTH/cable TV/satellite TV\n"
                    "- Landline/fixed-line telephone\n"
                    "- Enterprise telecom (leased lines, VPN, MPLS, SLA)\n"
                    "- ANY billing, payment, refund, service quality, or customer care issue "
                    "related to any of the above\n\n"
                    "SEMANTIC REASONING RULES:\n"
                    "1. Focus on the USER'S INTENT, not just the words they used.\n"
                    "2. 'Money deducted' in a telecom context = telecom billing issue.\n"
                    "3. 'Service not working' in a telecom context = telecom service disruption.\n"
                    "4. Vague complaints ARE telecom if the user came through the telecom menu.\n"
                    "5. Only reject if the query is CLEARLY about a non-telecom industry.\n"
                    + context_block +
                    '\n\nRespond with ONLY this JSON (no extra text):\n'
                    '{"reasoning": "<one sentence about why>", "is_telecom": true/false}'
                )},
                {"role": "user", "content": query},
            ],
            temperature=0,
            max_tokens=120,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        result = json.loads(raw)
        return result.get("is_telecom", False)
    except Exception:
        return True if sector_name else False


def identify_subprocess(query: str, sector_key: str) -> str:
    sector = TELECOM_MENU[sector_key]
    subprocess_details = get_subprocess_details(sector_key)
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    f"You are a semantic complaint classifier for: {sector['name']}.\n"
                    f"Sector description: {sector.get('description', '')}\n\n"
                    "Below are the available subprocesses:\n\n"
                    f"{subprocess_details}\n\n"
                    "Analyze the user's complaint and determine which subprocess it belongs to.\n\n"
                    "Respond with ONLY this JSON:\n"
                    '{"reasoning": "<brief explanation>", "matched_subprocess": "<exact name>", "confidence": <0.0 to 1.0>}'
                )},
                {"role": "user", "content": query},
            ],
            temperature=0,
            max_tokens=200,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        result = json.loads(raw)
        return result.get("matched_subprocess", "General Inquiry")
    except Exception:
        return "General Inquiry"


def detect_greeting(text: str) -> bool:
    """Semantically determine whether a message is a greeting in any language."""
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    "Determine if the user's message is a greeting or salutation in ANY language or mixed language. "
                    "A greeting includes (but is not limited to): hello, hi, hey, hiya, howdy, good morning, "
                    "good afternoon, good evening, namaste, namaskar, salaam, assalamu alaikum, "
                    "bonjour, hola, ciao, salam, sat sri akal, vanakkam, adab, greetings, what's up, "
                    "yo, sup, hii, helo, hai, or informal/phonetic variants in any script. "
                    "Mixed-language greetings (e.g. 'hello aur kaise ho', 'hi there bhai') also count. "
                    'Respond with ONLY valid JSON: {"is_greeting": true} or {"is_greeting": false}'
                )},
                {"role": "user", "content": text},
            ],
            temperature=0,
            max_tokens=20,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        result = json.loads(raw)
        return bool(result.get("is_greeting", True))
    except Exception:
        return True   # fail-open: treat ambiguous input as a greeting


def detect_language(text: str) -> str:
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    "Detect the language of the following text. "
                    'Respond with ONLY: {"language": "<language_name>", "code": "<iso_code>"}'
                )},
                {"role": "user", "content": text},
            ],
            temperature=0,
            max_tokens=50,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        result = json.loads(raw)
        return result.get("language", "English")
    except Exception:
        return "English"


def generate_resolution(query, sector_name, subprocess_name, language):
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    f"You are an expert telecom customer support agent. The user has a complaint "
                    f"under the sector: '{sector_name}' and subprocess: '{subprocess_name}'.\n\n"
                    "Provide a helpful response in the following format:\n"
                    "1. Acknowledge the issue empathetically\n"
                    "2. Provide 4-6 clear, actionable self-help troubleshooting steps\n"
                    "3. If the steps don't resolve the issue, advise contacting customer care\n"
                    "4. Provide a brief note about escalation options\n\n"
                    f"IMPORTANT: Respond entirely in {language}. "
                    "Keep the tone professional, empathetic, and helpful."
                )},
                {"role": "user", "content": query},
            ],
            temperature=0.4,
            max_tokens=1000,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"I apologize, but I encountered an error. Please try again. Error: {str(e)}"


def generate_single_solution(sector_name, subprocess_name, language, user_query="", previous_solutions=None, attempt=1):
    """Generate a single focused solution. If user_query is provided, tailor to it. Avoids repeating previous solutions."""
    prev_block = ""
    if previous_solutions:
        prev_block = (
            "\n\nIMPORTANT: The following solutions have ALREADY been provided and did NOT work. "
            "Do NOT repeat them. Provide a DIFFERENT approach:\n"
            + "\n---\n".join(previous_solutions)
        )

    query_block = ""
    if user_query:
        query_block = f"\n\nThe user described their specific issue as: \"{user_query}\""

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    f"You are an expert telecom customer support agent. The user has an issue "
                    f"under the sector: '{sector_name}' and subprocess: '{subprocess_name}'.\n\n"
                    f"This is solution attempt #{attempt} of 5.\n\n"
                    "Provide ONE focused, actionable solution with 2-3 clear steps. "
                    "Be concise and specific. Do not provide multiple alternative solutions — just one.\n"
                    "Acknowledge the issue briefly and give the steps."
                    + query_block
                    + prev_block +
                    f"\n\nIMPORTANT: Respond entirely in {language}. "
                    "Keep the tone professional, empathetic, and helpful."
                )},
                {"role": "user", "content": user_query if user_query else f"I have an issue with {subprocess_name} in {sector_name}"},
            ],
            temperature=0.5,
            max_tokens=500,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"I apologize, but I encountered an error. Please try again. Error: {str(e)}"


def translate_text(text: str, target_language: str) -> str:
    if target_language.lower() in ("english", "en"):
        return text
    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": f"Translate the following text to {target_language}. Keep formatting intact. Return ONLY the translation."},
                {"role": "user", "content": text},
            ],
            temperature=0,
            max_tokens=500,
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return text


def generate_chat_summary(messages_list, sector_name, subprocess_name):
    """Generate a summary of the chat conversation."""
    try:
        conversation = "\n".join([f"{m['sender']}: {m['content']}" for m in messages_list])
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    "Summarize this telecom support chat in 3-4 sentences. "
                    f"Category: {sector_name} > {subprocess_name}. "
                    "Include: what the issue was, what resolution was provided, and the outcome."
                )},
                {"role": "user", "content": conversation},
            ],
            temperature=0.3,
            max_tokens=200,
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return f"Chat about {sector_name} - {subprocess_name}. Customer query handled."


def generate_ref_number():
    ts = hex(int(time.time()))[2:].upper()
    rand = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"TC-{ts}-{rand}"


def auto_assign_priority(query_text, subprocess_name):
    """Simple priority assignment based on keywords."""
    text = (query_text + " " + subprocess_name).lower()
    if any(w in text for w in ["urgent", "critical", "emergency", "business down", "sla breach", "escalat"]):
        return "critical"
    if any(w in text for w in ["not working", "failed", "no signal", "dead", "down", "outage"]):
        return "high"
    if any(w in text for w in ["slow", "intermittent", "billing", "wrong charge", "refund"]):
        return "medium"
    return "low"



# AUTH ROUTES


@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.json
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")

    if not name or not email or not password:
        return jsonify({"error": "Name, email, and password are required"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already registered"}), 409

    user = User(name=name, email=email, role="customer")
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    token = create_access_token(identity=str(user.id))
    return jsonify({"token": token, "user": user.to_dict()}), 201


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.json
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")

    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify({"error": "Invalid email or password"}), 401

    token = create_access_token(identity=str(user.id))
    return jsonify({"token": token, "user": user.to_dict()})


@app.route("/api/auth/me", methods=["GET"])
@jwt_required()
def get_me():
    user = User.query.get(int(get_jwt_identity()))
    if not user:
        return jsonify({"error": "User not found"}), 404
    return jsonify({"user": user.to_dict()})



# CHATBOT ROUTES 
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/menu", methods=["GET"])
def get_menu():
    menu = {}
    for key, sector in TELECOM_MENU.items():
        menu[key] = {"name": sector["name"], "icon": sector["icon"]}
    return jsonify({"menu": menu})


@app.route("/api/subprocesses", methods=["POST"])
def get_subprocesses():
    data = request.json
    sector_key = data.get("sector_key")
    language = data.get("language", "English")
    if sector_key not in TELECOM_MENU:
        return jsonify({"error": "Invalid sector"}), 400
    sector = TELECOM_MENU[sector_key]
    subprocesses = {}
    for k, v in sector["subprocesses"].items():
        subprocesses[k] = v["name"] if isinstance(v, dict) else v
    if language.lower() not in ("english", "en"):
        translated = {}
        for k, v in subprocesses.items():
            translated[k] = translate_text(v, language)
        subprocesses = translated
    return jsonify({"sector_name": sector["name"], "subprocesses": subprocesses})


@app.route("/api/resolve", methods=["POST"])
def resolve_complaint():
    data = request.json
    query = data.get("query", "").strip()
    sector_key = data.get("sector_key")
    subprocess_key = data.get("subprocess_key")
    language = data.get("language", "English")
    if not query:
        return jsonify({"error": "Please enter your complaint/query."}), 400
    sector = TELECOM_MENU.get(sector_key, {})
    sector_name = sector.get("name", "Telecom")
    subprocess_name = get_subprocess_name(sector_key, subprocess_key)
    if not is_telecom_related(query, sector_name=sector_name, subprocess_name=subprocess_name):
        msg = (
            "I'm sorry, but I can only assist with **telecom-related** complaints. "
            "Your query doesn't appear to be telecom-related. Please try again."
        )
        translated_msg = translate_text(msg, language)
        return jsonify({"resolution": translated_msg, "is_telecom": False})
    if subprocess_name == "Others":
        subprocess_name = identify_subprocess(query, sector_key)
    resolution = generate_resolution(query, sector_name, subprocess_name, language)
    return jsonify({
        "resolution": resolution,
        "is_telecom": True,
        "identified_subprocess": subprocess_name,
    })


@app.route("/api/resolve-step", methods=["POST"])
def resolve_step():
    """Generate a single solution step. Used in the iterative resolution flow."""
    data = request.json
    sector_key = data.get("sector_key")
    subprocess_key = data.get("subprocess_key")
    user_query = data.get("query", "").strip()
    language = data.get("language", "English")
    previous_solutions = data.get("previous_solutions", [])
    attempt = data.get("attempt", 1)

    sector = TELECOM_MENU.get(sector_key, {})
    sector_name = sector.get("name", "Telecom")
    subprocess_name = get_subprocess_name(sector_key, subprocess_key)

    # If user provided a query, check if it's telecom-related
    if user_query:
        if not is_telecom_related(user_query, sector_name=sector_name, subprocess_name=subprocess_name):
            msg = (
                "I'm sorry, but I can only assist with **telecom-related** complaints. "
                "Your query doesn't appear to be telecom-related. Please try again."
            )
            translated_msg = translate_text(msg, language)
            return jsonify({"resolution": translated_msg, "is_telecom": False})

    solution = generate_single_solution(
        sector_name, subprocess_name, language,
        user_query=user_query,
        previous_solutions=previous_solutions,
        attempt=attempt,
    )
    return jsonify({
        "resolution": solution,
        "is_telecom": True,
        "attempt": attempt,
    })


@app.route("/api/detect-language", methods=["POST"])
def detect_lang():
    data = request.json
    text = data.get("text", "")
    language = detect_language(text)
    return jsonify({"language": language})


@app.route("/api/detect-greeting", methods=["POST"])
def detect_greeting_route():
    data = request.json
    text = data.get("text", "")
    is_greeting = detect_greeting(text)
    return jsonify({"is_greeting": is_greeting})


# ═══════════════════════════════════════════════════════════════════════════════
# CHAT SESSION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/chat/session", methods=["POST"])
@jwt_required()
def create_chat_session():
    user_id = int(get_jwt_identity())
    session = ChatSession(user_id=user_id, status="active")
    db.session.add(session)
    db.session.commit()
    return jsonify({"session": session.to_dict()}), 201


@app.route("/api/chat/session/<int:session_id>/message", methods=["POST"])
@jwt_required()
def add_chat_message(session_id):
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    data = request.json
    msg = ChatMessage(
        session_id=session_id,
        sender=data.get("sender", "user"),
        content=data.get("content", ""),
    )
    db.session.add(msg)

    # Update session metadata
    if data.get("sector_name"):
        session.sector_name = data["sector_name"]
    if data.get("subprocess_name"):
        session.subprocess_name = data["subprocess_name"]
    if data.get("query_text"):
        session.query_text = data["query_text"]
    if data.get("resolution"):
        session.resolution = data["resolution"]
    if data.get("language"):
        session.language = data["language"]

    db.session.commit()
    return jsonify({"message": msg.to_dict()})


@app.route("/api/chat/session/<int:session_id>/resolve", methods=["PUT"])
@jwt_required()
def resolve_session(session_id):
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    session.status = "resolved"
    session.resolved_at = datetime.now(timezone.utc)
    db.session.commit()

    # Generate summary after commit so resolve is never blocked
    try:
        msgs = [{"sender": m.sender, "content": m.content} for m in session.messages]
        session.summary = generate_chat_summary(msgs, session.sector_name, session.subprocess_name)
        db.session.commit()
    except Exception:
        pass

    return jsonify({"session": session.to_dict(), "summary": session.summary})


@app.route("/api/chat/session/<int:session_id>/escalate", methods=["PUT"])
@jwt_required()
def escalate_session(session_id):
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    session.status = "escalated"

    # Generate summary
    msgs = [{"sender": m.sender, "content": m.content} for m in session.messages]
    session.summary = generate_chat_summary(msgs, session.sector_name, session.subprocess_name)

    # Create ticket
    ref = generate_ref_number()
    priority = auto_assign_priority(session.query_text, session.subprocess_name)
    ticket = Ticket(
        chat_session_id=session_id,
        user_id=user_id,
        reference_number=ref,
        category=session.sector_name,
        subcategory=session.subprocess_name,
        description=session.query_text,
        status="pending",
        priority=priority,
    )
    db.session.add(ticket)
    db.session.commit()
    return jsonify({"session": session.to_dict(), "ticket": ticket.to_dict()})


@app.route("/api/chat/session/<int:session_id>/send-summary-email", methods=["POST"])
@jwt_required()
def send_summary_email(session_id):
    """Send chat summary to the user's email saved in DB."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404

    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403

    if not session.summary:
        return jsonify({"error": "No summary available for this session"}), 400

    # Build email HTML
    html_body = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #00338d 0%, #004fc4 100%); padding: 24px 30px; text-align: center;">
            <h1 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 600;">Customer Handling</h1>
            <p style="color: rgba(255,255,255,0.8); margin: 4px 0 0; font-size: 13px;">Chat Summary Report</p>
        </div>

        <!-- Body -->
        <div style="padding: 30px;">
            <p style="color: #1e293b; font-size: 15px; margin: 0 0 20px;">Hello <strong>{user.name}</strong>,</p>
            <p style="color: #64748b; font-size: 14px; line-height: 1.6; margin: 0 0 24px;">
                Here is the summary of your recent support chat session:
            </p>

            <!-- Session Details -->
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 20px; margin-bottom: 20px;">
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8; width: 130px;">Session ID</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">#{session.id}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Category</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.sector_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Issue Type</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.subprocess_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Status</td>
                        <td style="padding: 8px 0;">
                            <span style="background: {'#ecfdf5' if session.status == 'resolved' else '#fef3c7'}; color: {'#047857' if session.status == 'resolved' else '#b45309'}; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600;">
                                {session.status.upper()}
                            </span>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Language</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.language or 'English'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Date</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.created_at.strftime('%B %d, %Y at %I:%M %p') if session.created_at else 'N/A'}</td>
                    </tr>
                </table>
            </div>

            <!-- Summary -->
            <div style="border-left: 3px solid #10b981; background: #f0fdf4; border-radius: 0 10px 10px 0; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #047857; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 10px;">Chat Summary</h3>
                <p style="color: #1e293b; font-size: 14px; line-height: 1.7; margin: 0;">{session.summary}</p>
            </div>

            <!-- Your Query -->
            {f'''
            <div style="background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #2563eb; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 8px;">Your Query</h3>
                <p style="color: #1e293b; font-size: 14px; line-height: 1.6; margin: 0;">{session.query_text}</p>
            </div>
            ''' if session.query_text else ''}

            <!-- Ticket Info -->
            {f'''
            <div style="background: #fffbeb; border: 1px solid #fde68a; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #b45309; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 8px;">Escalation Ticket</h3>
                <p style="color: #1e293b; font-size: 14px; margin: 0;">Reference: <strong>{session.ticket.reference_number}</strong></p>
            </div>
            ''' if session.ticket else ''}

            <p style="color: #64748b; font-size: 13px; line-height: 1.6; margin: 20px 0 0;">
                If you have further questions, feel free to start a new chat session anytime.
            </p>
        </div>

        <!-- Footer -->
        <div style="background: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px 30px; text-align: center;">
            <p style="color: #94a3b8; font-size: 12px; margin: 0;">Customer Handling &mdash; AI-Powered Support</p>
        </div>
    </div>
    """

    try:
        msg = Message(
            subject=f"Chat Summary - {session.sector_name or 'Telecom Support'} (Session #{session.id})",
            recipients=[user.email],
            html=html_body,
        )
        mail.send(msg)
        return jsonify({"message": f"Summary sent to {user.email}"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to send email: {str(e)}"}), 500


@app.route("/api/chat/session/<int:session_id>", methods=["GET"])
@jwt_required()
def get_chat_session(session_id):
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    return jsonify({
        "session": session.to_dict(),
        "messages": [m.to_dict() for m in session.messages],
    })


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOMER ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/customer/dashboard", methods=["GET"])
@jwt_required()
def customer_dashboard():
    user_id = int(get_jwt_identity())
    total = ChatSession.query.filter_by(user_id=user_id).count()
    resolved = ChatSession.query.filter_by(user_id=user_id, status="resolved").count()
    escalated = ChatSession.query.filter_by(user_id=user_id, status="escalated").count()
    active = ChatSession.query.filter_by(user_id=user_id, status="active").count()
    pending_tickets = Ticket.query.filter_by(user_id=user_id).filter(
        Ticket.status.in_(["pending", "in_progress"])
    ).count()

    recent_sessions = ChatSession.query.filter_by(user_id=user_id).order_by(
        ChatSession.created_at.desc()
    ).all()

    return jsonify({
        "stats": {
            "total_chats": total,
            "resolved": resolved,
            "escalated": escalated,
            "active": active,
            "pending_tickets": pending_tickets,
        },
        "recent_sessions": [s.to_dict() for s in recent_sessions],
    })


@app.route("/api/customer/active-session", methods=["GET"])
@jwt_required()
def customer_active_session():
    """Return the most recent active chat session for the current user, with messages."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.filter_by(user_id=user_id, status="active").order_by(
        ChatSession.created_at.desc()
    ).first()
    if not session:
        return jsonify({"session": None, "messages": []})
    return jsonify({
        "session": session.to_dict(),
        "messages": [m.to_dict() for m in session.messages],
    })


@app.route("/api/customer/pending-feedback", methods=["GET"])
@jwt_required()
def customer_pending_feedback():
    """Return resolved/escalated sessions that the user hasn't given feedback for."""
    user_id = int(get_jwt_identity())
    # Subquery: session IDs that already have feedback from this user
    feedback_session_ids = db.session.query(Feedback.chat_session_id).filter(
        Feedback.user_id == user_id,
        Feedback.chat_session_id.isnot(None),
    ).subquery()

    sessions = ChatSession.query.filter(
        ChatSession.user_id == user_id,
        ChatSession.status.in_(["resolved", "escalated"]),
        ~ChatSession.id.in_(feedback_session_ids),
    ).order_by(ChatSession.created_at.desc()).all()

    return jsonify({
        "sessions": [s.to_dict() for s in sessions],
    })


@app.route("/api/customer/sessions", methods=["GET"])
@jwt_required()
def customer_sessions():
    user_id = int(get_jwt_identity())
    sessions = ChatSession.query.filter_by(user_id=user_id).order_by(
        ChatSession.created_at.desc()
    ).all()
    return jsonify({"sessions": [s.to_dict() for s in sessions]})


@app.route("/api/customer/tickets", methods=["GET"])
@jwt_required()
def customer_tickets():
    user_id = int(get_jwt_identity())
    tickets = Ticket.query.filter_by(user_id=user_id).order_by(
        Ticket.created_at.desc()
    ).all()
    return jsonify({"tickets": [t.to_dict() for t in tickets]})


# ═══════════════════════════════════════════════════════════════════════════════
# FEEDBACK ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/feedback", methods=["POST"])
@jwt_required()
def submit_feedback():
    user_id = int(get_jwt_identity())
    data = request.json
    fb = Feedback(
        user_id=user_id,
        chat_session_id=data.get("chat_session_id"),
        rating=data.get("rating", 0),
        comment=data.get("comment", ""),
    )
    db.session.add(fb)
    db.session.commit()
    return jsonify({"feedback": fb.to_dict()}), 201


@app.route("/api/feedback/list", methods=["GET"])
@jwt_required()
def list_feedback():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role == "customer":
        feedbacks = Feedback.query.filter_by(user_id=user_id).order_by(Feedback.created_at.desc()).all()
    else:
        feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    return jsonify({"feedbacks": [f.to_dict() for f in feedbacks]})


# ═══════════════════════════════════════════════════════════════════════════════
# MANAGER / CTO ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/manager/dashboard", methods=["GET"])
@jwt_required()
def manager_dashboard():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    total_chats = ChatSession.query.count()
    resolved_chats = ChatSession.query.filter_by(status="resolved").count()
    escalated_chats = ChatSession.query.filter_by(status="escalated").count()
    active_chats = ChatSession.query.filter_by(status="active").count()

    total_tickets = Ticket.query.count()
    pending_tickets = Ticket.query.filter_by(status="pending").count()
    in_progress_tickets = Ticket.query.filter_by(status="in_progress").count()
    resolved_tickets = Ticket.query.filter_by(status="resolved").count()
    escalated_tickets = Ticket.query.filter_by(status="escalated").count()

    critical_tickets = Ticket.query.filter_by(priority="critical", status="pending").count()
    high_tickets = Ticket.query.filter_by(priority="high", status="pending").count()

    total_feedback = Feedback.query.count()
    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(Feedback.rating > 0).scalar() or 0
    satisfied_count = Feedback.query.filter(Feedback.rating >= 4).count()
    csat_score = round((satisfied_count / max(total_feedback, 1)) * 100, 1)

    total_users = User.query.filter_by(role="customer").count()

    # Category breakdown
    categories = db.session.query(
        ChatSession.sector_name, db.func.count(ChatSession.id)
    ).group_by(ChatSession.sector_name).all()

    return jsonify({
        "stats": {
            "total_chats": total_chats,
            "resolved_chats": resolved_chats,
            "escalated_chats": escalated_chats,
            "active_chats": active_chats,
            "total_tickets": total_tickets,
            "pending_tickets": pending_tickets,
            "in_progress_tickets": in_progress_tickets,
            "resolved_tickets": resolved_tickets,
            "escalated_tickets": escalated_tickets,
            "critical_tickets": critical_tickets,
            "high_tickets": high_tickets,
            "total_feedback": total_feedback,
            "avg_rating": round(float(avg_rating), 1),
            "csat_score": csat_score,
            "total_customers": total_users,
        },
        "category_breakdown": [{"name": c[0] or "Unknown", "count": c[1]} for c in categories],
    })


@app.route("/api/manager/tickets", methods=["GET"])
@jwt_required()
def manager_tickets():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    priority = request.args.get("priority")
    category = request.args.get("category")
    search = request.args.get("search")

    query = Ticket.query
    if status:
        query = query.filter_by(status=status)
    if priority:
        query = query.filter_by(priority=priority)
    if category:
        query = query.filter_by(category=category)
    if search:
        query = query.join(User, Ticket.user_id == User.id).filter(
            db.or_(
                User.name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
                Ticket.reference_number.ilike(f"%{search}%"),
                Ticket.description.ilike(f"%{search}%"),
            )
        )

    tickets = query.order_by(Ticket.created_at.desc()).all()
    return jsonify({"tickets": [t.to_dict() for t in tickets]})


@app.route("/api/manager/tickets/<int:ticket_id>", methods=["PUT"])
@jwt_required()
def update_ticket(ticket_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404

    data = request.json
    if "status" in data:
        ticket.status = data["status"]
        if data["status"] == "resolved":
            ticket.resolved_at = datetime.now(timezone.utc)
    if "priority" in data:
        ticket.priority = data["priority"]
    if "assigned_to" in data:
        ticket.assigned_to = data["assigned_to"]
    if "resolution_notes" in data:
        ticket.resolution_notes = data["resolution_notes"]

    db.session.commit()
    return jsonify({"ticket": ticket.to_dict()})


@app.route("/api/manager/chats", methods=["GET"])
@jwt_required()
def manager_chats():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    query = ChatSession.query
    if status:
        query = query.filter_by(status=status)

    sessions = query.order_by(ChatSession.created_at.desc()).all()
    return jsonify({"sessions": [s.to_dict() for s in sessions]})


@app.route("/api/manager/users", methods=["GET"])
@jwt_required()
def manager_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    managers = User.query.filter(User.role.in_(["manager"])).all()
    return jsonify({"managers": [u.to_dict() for u in managers]})


# ═══════════════════════════════════════════════════════════════════════════════
# CTO-SPECIFIC ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cto/overview", methods=["GET"])
@jwt_required()
def cto_overview():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "cto":
        return jsonify({"error": "Unauthorized"}), 403

    # Resolution rate
    total = ChatSession.query.count() or 1
    resolved = ChatSession.query.filter_by(status="resolved").count()
    resolution_rate = round((resolved / total) * 100, 1)

    # Avg rating
    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(Feedback.rating > 0).scalar() or 0

    # Tickets by priority
    priorities = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).group_by(Ticket.priority).all()

    # Monthly trends (last 6 months)
    six_months_ago = datetime.now(timezone.utc) - timedelta(days=180)
    monthly = db.session.query(
        db.func.date_trunc("month", ChatSession.created_at).label("month"),
        db.func.count(ChatSession.id),
    ).filter(ChatSession.created_at >= six_months_ago).group_by("month").order_by("month").all()

    return jsonify({
        "resolution_rate": resolution_rate,
        "avg_rating": round(float(avg_rating), 1),
        "total_customers": User.query.filter_by(role="customer").count(),
        "total_sessions": total,
        "priority_breakdown": [{"priority": p[0], "count": p[1]} for p in priorities],
        "monthly_trends": [{"month": m[0].isoformat() if m[0] else "", "count": m[1]} for m in monthly],
    })


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE ID GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

ROLE_PREFIX = {
    "manager": "MGR",
    "human_agent": "HA",
    "cto": "CTO",
    "admin": "ADM",
}


def generate_employee_id(role):
    prefix = ROLE_PREFIX.get(role)
    if not prefix:
        return None
    existing = User.query.filter(User.employee_id.like(f"{prefix}%")).order_by(User.employee_id.desc()).first()
    if existing and existing.employee_id:
        num = int(existing.employee_id[len(prefix):]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/admin/dashboard", methods=["GET"])
@jwt_required()
def admin_dashboard():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    # User counts by role
    user_counts = db.session.query(
        User.role, db.func.count(User.id)
    ).group_by(User.role).all()

    total_users = sum(c[1] for c in user_counts)

    # Chat stats
    total_chats = ChatSession.query.count()
    resolved_chats = ChatSession.query.filter_by(status="resolved").count()
    escalated_chats = ChatSession.query.filter_by(status="escalated").count()
    active_chats = ChatSession.query.filter_by(status="active").count()

    # Ticket stats
    total_tickets = Ticket.query.count()
    pending_tickets = Ticket.query.filter_by(status="pending").count()
    in_progress_tickets = Ticket.query.filter_by(status="in_progress").count()
    resolved_tickets = Ticket.query.filter_by(status="resolved").count()
    critical_tickets = Ticket.query.filter_by(priority="critical").count()
    high_tickets = Ticket.query.filter_by(priority="high").count()

    # Feedback
    total_feedback = Feedback.query.count()
    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(Feedback.rating > 0).scalar() or 0
    satisfied_count = Feedback.query.filter(Feedback.rating >= 4).count()
    csat_score = round((satisfied_count / max(total_feedback, 1)) * 100, 1)

    # Resolution rate
    resolution_rate = round((resolved_chats / max(total_chats, 1)) * 100, 1)

    # Category breakdown
    categories = db.session.query(
        ChatSession.sector_name, db.func.count(ChatSession.id)
    ).group_by(ChatSession.sector_name).all()

    return jsonify({
        "stats": {
            "total_users": total_users,
            "total_chats": total_chats,
            "resolved_chats": resolved_chats,
            "escalated_chats": escalated_chats,
            "active_chats": active_chats,
            "total_tickets": total_tickets,
            "pending_tickets": pending_tickets,
            "in_progress_tickets": in_progress_tickets,
            "resolved_tickets": resolved_tickets,
            "critical_tickets": critical_tickets,
            "high_tickets": high_tickets,
            "total_feedback": total_feedback,
            "avg_rating": round(float(avg_rating), 1),
            "csat_score": csat_score,
            "resolution_rate": resolution_rate,
        },
        "user_breakdown": [{"role": r[0], "count": r[1]} for r in user_counts],
        "category_breakdown": [{"name": c[0] or "Unknown", "count": c[1]} for c in categories],
    })


@app.route("/api/admin/users", methods=["GET"])
@jwt_required()
def admin_list_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    role_filter = request.args.get("role")
    search = request.args.get("search")

    query = User.query
    if role_filter:
        query = query.filter_by(role=role_filter)
    if search:
        query = query.filter(
            db.or_(
                User.name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
            )
        )

    users = query.order_by(User.created_at.desc()).all()
    return jsonify({"users": [u.to_dict() for u in users]})


@app.route("/api/admin/users", methods=["POST"])
@jwt_required()
def admin_create_user():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    role = data.get("role", "customer").lower()

    if not name or not email or not password:
        return jsonify({"error": "Name, email, and password are required"}), 400
    if role not in ("customer", "manager", "human_agent", "cto", "admin"):
        return jsonify({"error": "Invalid role"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already registered"}), 409

    emp_id = generate_employee_id(role)
    new_user = User(name=name, email=email, role=role, employee_id=emp_id)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"user": new_user.to_dict()}), 201


@app.route("/api/admin/users/upload", methods=["POST"])
@jwt_required()
def admin_upload_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only Excel files (.xlsx, .xls) are allowed"}), 400

    from openpyxl import load_workbook
    import io

    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
    except Exception:
        return jsonify({"error": "Could not read Excel file"}), 400

    # Parse headers from first row
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    required = {"name", "email", "role"}
    header_set = set(headers)
    if not required.issubset(header_set):
        missing = required - header_set
        return jsonify({"error": f"Missing required columns: {', '.join(missing)}. Required: Name, Email, Role"}), 400

    col_map = {h: i for i, h in enumerate(headers)}
    valid_roles = {"manager", "human_agent", "cto", "admin"}
    created = 0
    updated = 0
    skipped = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[col_map["name"]] or "").strip()
        email = str(row[col_map["email"]] or "").strip().lower()
        role = str(row[col_map["role"]] or "").strip().lower().replace(" ", "_")

        if not name or not email:
            skipped.append(f"Row {row_num}: missing name or email")
            continue
        if role not in valid_roles:
            skipped.append(f"Row {row_num}: invalid role '{role}' for {email}")
            continue

        # Check for employee_id column
        emp_id_from_excel = None
        if "employee id" in col_map:
            emp_id_from_excel = str(row[col_map["employee id"]] or "").strip() or None
        elif "employee_id" in col_map:
            emp_id_from_excel = str(row[col_map["employee_id"]] or "").strip() or None

        existing = User.query.filter_by(email=email).first()
        if existing:
            existing.name = name
            existing.role = role
            if emp_id_from_excel:
                existing.employee_id = emp_id_from_excel
            elif not existing.employee_id:
                existing.employee_id = generate_employee_id(role)
            updated += 1
        else:
            emp_id = emp_id_from_excel or generate_employee_id(role)
            new_user = User(name=name, email=email, role=role, employee_id=emp_id)
            new_user.set_password("Welcome@123")
            db.session.add(new_user)
            created += 1

    db.session.commit()
    return jsonify({
        "message": f"Upload complete: {created} created, {updated} updated",
        "created": created,
        "updated": updated,
        "skipped": skipped,
    })


@app.route("/api/admin/users/<int:uid>", methods=["PUT"])
@jwt_required()
def admin_update_user(uid):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    target = User.query.get(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404

    data = request.json
    if "name" in data:
        target.name = data["name"].strip()
    if "email" in data:
        new_email = data["email"].strip().lower()
        existing = User.query.filter_by(email=new_email).first()
        if existing and existing.id != uid:
            return jsonify({"error": "Email already in use"}), 409
        target.email = new_email
    if "role" in data:
        new_role = data["role"].lower()
        if new_role not in ("customer", "manager", "human_agent", "cto", "admin"):
            return jsonify({"error": "Invalid role"}), 400
        if uid == user_id and new_role != "admin":
            return jsonify({"error": "Cannot change your own role"}), 400
        target.role = new_role
        if new_role == "customer":
            target.employee_id = None
        else:
            target.employee_id = generate_employee_id(new_role)
    if "password" in data and data["password"]:
        target.set_password(data["password"])

    db.session.commit()
    return jsonify({"user": target.to_dict()})


@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@jwt_required()
def admin_delete_user(uid):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if uid == user_id:
        return jsonify({"error": "Cannot delete your own account"}), 400

    target = User.query.get(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404

    # Delete associated data
    Feedback.query.filter_by(user_id=uid).delete()
    ChatMessage.query.filter(
        ChatMessage.session_id.in_(
            db.session.query(ChatSession.id).filter_by(user_id=uid)
        )
    ).delete(synchronize_session=False)
    Ticket.query.filter_by(user_id=uid).delete()
    ChatSession.query.filter_by(user_id=uid).delete()
    db.session.delete(target)
    db.session.commit()
    return jsonify({"message": "User deleted"})


@app.route("/api/admin/agent-tickets", methods=["GET"])
@jwt_required()
def admin_agent_tickets():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    agent_id = request.args.get("agent_id")
    search = request.args.get("search")

    # Alias User for the assignee join
    AgentUser = db.aliased(User)
    query = (
        Ticket.query
        .join(AgentUser, Ticket.assigned_to == AgentUser.id)
        .filter(AgentUser.role == "human_agent")
    )

    if status:
        query = query.filter(Ticket.status == status)
    if agent_id:
        query = query.filter(Ticket.assigned_to == int(agent_id))
    if search:
        CustomerUser = db.aliased(User)
        query = (
            query
            .join(CustomerUser, Ticket.user_id == CustomerUser.id)
            .filter(db.or_(
                CustomerUser.name.ilike(f"%{search}%"),
                CustomerUser.email.ilike(f"%{search}%"),
                Ticket.reference_number.ilike(f"%{search}%"),
            ))
        )

    tickets = query.order_by(Ticket.created_at.desc()).all()
    agents = User.query.filter_by(role="human_agent").order_by(User.name).all()

    return jsonify({
        "tickets": [t.to_dict() for t in tickets],
        "agents": [{"id": a.id, "name": a.name} for a in agents],
    })


@app.route("/api/admin/feedback", methods=["GET"])
@jwt_required()
def admin_feedback():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    result = []
    for f in feedbacks:
        fd = f.to_dict()
        if f.chat_session:
            fd["session_sector"] = f.chat_session.sector_name
            fd["session_subprocess"] = f.chat_session.subprocess_name
        result.append(fd)
    return jsonify({"feedbacks": result})


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS & ANALYTICS ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

SLA_DEFAULTS = {
    "sla_critical": {"value": "4", "description": "SLA target hours for Critical priority"},
    "sla_high": {"value": "12", "description": "SLA target hours for High priority"},
    "sla_medium": {"value": "48", "description": "SLA target hours for Medium priority"},
    "sla_low": {"value": "120", "description": "SLA target hours for Low priority"},
}


def get_sla_targets():
    targets = {}
    for key in ["sla_critical", "sla_high", "sla_medium", "sla_low"]:
        setting = SystemSetting.query.filter_by(key=key).first()
        priority = key.replace("sla_", "")
        targets[priority] = float(setting.value) if setting else float(SLA_DEFAULTS[key]["value"])
    return targets


def get_date_range(range_param):
    now = datetime.now(timezone.utc)
    if range_param == "7d":
        return now - timedelta(days=7)
    elif range_param == "90d":
        return now - timedelta(days=90)
    elif range_param == "12m":
        return now - timedelta(days=365)
    else:
        return now - timedelta(days=30)


def get_previous_period(range_param):
    now = datetime.now(timezone.utc)
    if range_param == "7d":
        return now - timedelta(days=14), now - timedelta(days=7)
    elif range_param == "90d":
        return now - timedelta(days=180), now - timedelta(days=90)
    elif range_param == "12m":
        return now - timedelta(days=730), now - timedelta(days=365)
    else:
        return now - timedelta(days=60), now - timedelta(days=30)


def calc_trend(current, previous):
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 1)


@app.route("/api/reports/overview", methods=["GET"])
@jwt_required()
def reports_overview():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)

    # Current period tickets
    current_tickets = Ticket.query.filter(Ticket.created_at >= start_date)
    resolved_current = current_tickets.filter(Ticket.status == "resolved").count()

    # Previous period
    prev_tickets = Ticket.query.filter(Ticket.created_at >= prev_start, Ticket.created_at < prev_end)
    resolved_prev = prev_tickets.filter(Ticket.status == "resolved").count()

    # Avg resolution time (current period)
    resolved_with_time = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()
    if resolved_with_time:
        total_hours = sum(
            (t.resolved_at - t.created_at).total_seconds() / 3600
            for t in resolved_with_time
        )
        avg_resolution = round(total_hours / len(resolved_with_time), 1)
    else:
        avg_resolution = 0

    # Previous avg resolution
    prev_resolved_with_time = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    if prev_resolved_with_time:
        prev_total_hours = sum(
            (t.resolved_at - t.created_at).total_seconds() / 3600
            for t in prev_resolved_with_time
        )
        prev_avg_resolution = round(prev_total_hours / len(prev_resolved_with_time), 1)
    else:
        prev_avg_resolution = 0

    # CSAT
    current_feedback = Feedback.query.filter(Feedback.created_at >= start_date)
    total_fb = current_feedback.count()
    satisfied = current_feedback.filter(Feedback.rating >= 4).count()
    csat = round((satisfied / max(total_fb, 1)) * 100, 1)

    prev_feedback = Feedback.query.filter(Feedback.created_at >= prev_start, Feedback.created_at < prev_end)
    prev_total_fb = prev_feedback.count()
    prev_satisfied = prev_feedback.filter(Feedback.rating >= 4).count()
    prev_csat = round((prev_satisfied / max(prev_total_fb, 1)) * 100, 1)

    # SLA compliance
    sla_targets = get_sla_targets()
    all_resolved = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()
    within_sla = 0
    for t in all_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            within_sla += 1
    sla_compliance = round((within_sla / max(len(all_resolved), 1)) * 100, 1)

    prev_all_resolved = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    prev_within_sla = 0
    for t in prev_all_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            prev_within_sla += 1
    prev_sla = round((prev_within_sla / max(len(prev_all_resolved), 1)) * 100, 1)

    # Resolution trends (monthly)
    resolution_trends = db.session.query(
        db.func.date_trunc("month", Ticket.resolved_at).label("month"),
        db.func.avg(
            db.func.extract("epoch", Ticket.resolved_at - Ticket.created_at) / 3600
        ).label("avg_hours"),
        db.func.count(Ticket.id).label("volume")
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.created_at >= start_date
    ).group_by("month").order_by("month").all()

    # Weekly volume
    weekly_volume = db.session.query(
        db.func.extract("dow", Ticket.created_at).label("dow"),
        db.func.count(Ticket.id).label("opened")
    ).filter(Ticket.created_at >= start_date).group_by("dow").order_by("dow").all()

    weekly_resolved = db.session.query(
        db.func.extract("dow", Ticket.resolved_at).label("dow"),
        db.func.count(Ticket.id).label("resolved")
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.resolved_at >= start_date
    ).group_by("dow").order_by("dow").all()

    day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    opened_map = {int(r[0]): r[1] for r in weekly_volume}
    resolved_map = {int(r[0]): r[1] for r in weekly_resolved}
    weekly_data = [
        {"day": day_names[i], "opened": opened_map.get(i, 0), "resolved": resolved_map.get(i, 0)}
        for i in range(7)
    ]

    # Category breakdown
    categories = db.session.query(
        Ticket.category, db.func.count(Ticket.id)
    ).filter(Ticket.created_at >= start_date).group_by(Ticket.category).all()

    # Priority distribution
    priorities = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).filter(Ticket.created_at >= start_date).group_by(Ticket.priority).all()

    return jsonify({
        "total_resolved": resolved_current,
        "resolved_trend": calc_trend(resolved_current, resolved_prev),
        "avg_resolution_hours": avg_resolution,
        "resolution_trend": calc_trend(avg_resolution, prev_avg_resolution),
        "csat_score": csat,
        "csat_trend": calc_trend(csat, prev_csat),
        "sla_compliance": sla_compliance,
        "sla_trend": calc_trend(sla_compliance, prev_sla),
        "resolution_trends": [
            {
                "month": r[0].strftime("%b %Y") if r[0] else "",
                "avg_hours": round(float(r[1] or 0), 1),
                "volume": r[2]
            } for r in resolution_trends
        ],
        "weekly_volume": weekly_data,
        "category_breakdown": [{"name": c[0] or "Other", "count": c[1]} for c in categories],
        "priority_distribution": [{"priority": p[0], "count": p[1]} for p in priorities],
    })


@app.route("/api/reports/agents", methods=["GET"])
@jwt_required()
def reports_agents():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)

    managers = User.query.filter(User.role.in_(["manager"])).all()
    agents_data = []

    for mgr in managers:
        assigned = Ticket.query.filter(
            Ticket.assigned_to == mgr.id,
            Ticket.created_at >= start_date
        )
        resolved = assigned.filter(Ticket.status == "resolved").count()
        pending = assigned.filter(Ticket.status.in_(["pending", "in_progress"])).count()
        escalated = assigned.filter(Ticket.status == "escalated").count()

        resolved_tickets = Ticket.query.filter(
            Ticket.assigned_to == mgr.id,
            Ticket.status == "resolved",
            Ticket.resolved_at.isnot(None),
            Ticket.created_at >= start_date
        ).all()

        if resolved_tickets:
            avg_time = round(sum(
                (t.resolved_at - t.created_at).total_seconds() / 3600
                for t in resolved_tickets
            ) / len(resolved_tickets), 1)
        else:
            avg_time = 0

        agent_feedback = db.session.query(db.func.avg(Feedback.rating)).join(
            Ticket, Feedback.chat_session_id == Ticket.chat_session_id
        ).filter(
            Ticket.assigned_to == mgr.id,
            Feedback.rating > 0,
            Feedback.created_at >= start_date
        ).scalar()

        agents_data.append({
            "id": mgr.id,
            "name": mgr.name,
            "resolved": resolved,
            "pending": pending,
            "escalated": escalated,
            "avg_resolution_hours": avg_time,
            "avg_rating": round(float(agent_feedback or 0), 1),
        })

    top_performer = max(agents_data, key=lambda x: x["resolved"], default=None)
    fastest = min(
        [a for a in agents_data if a["avg_resolution_hours"] > 0],
        key=lambda x: x["avg_resolution_hours"], default=None
    )
    highest_rated = max(
        [a for a in agents_data if a["avg_rating"] > 0],
        key=lambda x: x["avg_rating"], default=None
    )

    return jsonify({
        "agents": agents_data,
        "total_agents": len(managers),
        "top_performer": {"name": top_performer["name"], "resolved": top_performer["resolved"]} if top_performer else None,
        "fastest_agent": {"name": fastest["name"], "hours": fastest["avg_resolution_hours"]} if fastest else None,
        "highest_rated": {"name": highest_rated["name"], "rating": highest_rated["avg_rating"]} if highest_rated else None,
    })


@app.route("/api/reports/csat", methods=["GET"])
@jwt_required()
def reports_csat():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)

    # Current CSAT
    current_fb = Feedback.query.filter(Feedback.created_at >= start_date)
    total_responses = current_fb.count()
    satisfied = current_fb.filter(Feedback.rating >= 4).count()
    csat = round((satisfied / max(total_responses, 1)) * 100, 1)

    prev_fb = Feedback.query.filter(Feedback.created_at >= prev_start, Feedback.created_at < prev_end)
    prev_total = prev_fb.count()
    prev_satisfied = prev_fb.filter(Feedback.rating >= 4).count()
    prev_csat = round((prev_satisfied / max(prev_total, 1)) * 100, 1)

    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(
        Feedback.created_at >= start_date, Feedback.rating > 0
    ).scalar() or 0

    # Response rate
    resolved_tickets = Ticket.query.filter(
        Ticket.created_at >= start_date, Ticket.status == "resolved"
    ).count()
    response_rate = round((total_responses / max(resolved_tickets, 1)) * 100, 1)

    # Monthly CSAT trend
    monthly_csat = db.session.query(
        db.func.date_trunc("month", Feedback.created_at).label("month"),
        db.func.count(Feedback.id).label("total"),
        db.func.count(db.case((Feedback.rating >= 4, 1))).label("satisfied")
    ).filter(Feedback.created_at >= start_date).group_by("month").order_by("month").all()

    # Feedback distribution (1-5 stars)
    distribution = db.session.query(
        Feedback.rating, db.func.count(Feedback.id)
    ).filter(
        Feedback.created_at >= start_date, Feedback.rating > 0
    ).group_by(Feedback.rating).order_by(Feedback.rating).all()

    dist_map = {r[0]: r[1] for r in distribution}
    feedback_dist = [{"stars": i, "count": dist_map.get(i, 0)} for i in range(1, 6)]

    # Response volume trend
    volume_trend = db.session.query(
        db.func.date_trunc("month", Feedback.created_at).label("month"),
        db.func.count(Feedback.id).label("count")
    ).filter(Feedback.created_at >= start_date).group_by("month").order_by("month").all()

    return jsonify({
        "current_csat": csat,
        "csat_trend": calc_trend(csat, prev_csat),
        "total_responses": total_responses,
        "responses_trend": calc_trend(total_responses, prev_total),
        "avg_rating": round(float(avg_rating), 1),
        "response_rate": min(response_rate, 100),
        "csat_monthly": [
            {
                "month": m[0].strftime("%b %Y") if m[0] else "",
                "csat": round((m[2] / max(m[1], 1)) * 100, 1)
            } for m in monthly_csat
        ],
        "feedback_distribution": feedback_dist,
        "response_volume": [
            {"month": v[0].strftime("%b %Y") if v[0] else "", "count": v[1]}
            for v in volume_trend
        ],
    })


@app.route("/api/reports/sla", methods=["GET"])
@jwt_required()
def reports_sla():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)
    sla_targets = get_sla_targets()

    resolved = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()

    within = 0
    near_breach = 0
    breached = 0
    first_response_times = []

    priority_stats = {}
    for p in ["critical", "high", "medium", "low"]:
        priority_stats[p] = {"target": sla_targets.get(p, 48), "times": [], "within": 0, "breached": 0}

    for t in resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        first_response_times.append(hours)

        if t.priority in priority_stats:
            priority_stats[t.priority]["times"].append(hours)
            if hours <= target:
                priority_stats[t.priority]["within"] += 1
            else:
                priority_stats[t.priority]["breached"] += 1

        if hours <= target:
            within += 1
        elif hours <= target * 1.0 and hours > target * 0.8:
            near_breach += 1
        else:
            pct = hours / target if target > 0 else 999
            if pct > 1.0:
                breached += 1
            elif pct > 0.8:
                near_breach += 1
            else:
                within += 1

    total = max(len(resolved), 1)
    compliance_pct = round((within / total) * 100, 1)
    near_pct = round((near_breach / total) * 100, 1)
    breached_pct = round((breached / total) * 100, 1)
    avg_first_response = round(sum(first_response_times) / max(len(first_response_times), 1), 1)

    # Previous period compliance
    prev_resolved = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    prev_within = 0
    for t in prev_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            prev_within += 1
    prev_compliance = round((prev_within / max(len(prev_resolved), 1)) * 100, 1)

    # SLA targets with actual averages
    sla_target_list = []
    for p in ["critical", "high", "medium", "low"]:
        ps = priority_stats[p]
        avg_actual = round(sum(ps["times"]) / max(len(ps["times"]), 1), 1) if ps["times"] else 0
        sla_target_list.append({
            "priority": p,
            "target_hours": ps["target"],
            "actual_hours": avg_actual,
            "status": "within" if avg_actual <= ps["target"] else "breached",
            "total": len(ps["times"]),
        })

    # Monthly breach trend
    monthly_trend = db.session.query(
        db.func.date_trunc("month", Ticket.resolved_at).label("month"),
        Ticket.priority,
        Ticket.resolved_at,
        Ticket.created_at
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.created_at >= start_date
    ).all()

    month_data = {}
    for t in monthly_trend:
        month_key = t[0].strftime("%b %Y") if t[0] else "Unknown"
        if month_key not in month_data:
            month_data[month_key] = {"compliant": 0, "near_breach": 0, "breached": 0, "total": 0}
        hours = (t[2] - t[3]).total_seconds() / 3600
        target = sla_targets.get(t[1], 48)
        month_data[month_key]["total"] += 1
        pct_of_target = hours / target if target > 0 else 999
        if pct_of_target <= 0.8:
            month_data[month_key]["compliant"] += 1
        elif pct_of_target <= 1.0:
            month_data[month_key]["near_breach"] += 1
        else:
            month_data[month_key]["breached"] += 1

    breach_trend = []
    for month_key, data in sorted(month_data.items()):
        t = max(data["total"], 1)
        breach_trend.append({
            "month": month_key,
            "compliant": round((data["compliant"] / t) * 100, 1),
            "near_breach": round((data["near_breach"] / t) * 100, 1),
            "breached": round((data["breached"] / t) * 100, 1),
        })

    return jsonify({
        "compliance_percentage": compliance_pct,
        "compliance_trend": calc_trend(compliance_pct, prev_compliance),
        "near_breach_percentage": near_pct,
        "breached_percentage": breached_pct,
        "avg_first_response": avg_first_response,
        "sla_targets": sla_target_list,
        "breach_trend": breach_trend,
        "within_count": within,
        "near_breach_count": near_breach,
        "breached_count": breached,
    })


@app.route("/api/reports/export", methods=["GET"])
@jwt_required()
def reports_export():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    fmt = request.args.get("format", "csv")
    section = request.args.get("section", "overview")
    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)

    if fmt == "csv":
        import io
        import csv
        from flask import Response

        output = io.StringIO()
        writer = csv.writer(output)

        if section == "overview":
            tickets = Ticket.query.filter(Ticket.created_at >= start_date).all()
            writer.writerow(["Reference", "Category", "Priority", "Status", "Created", "Resolved", "Resolution Hours"])
            for t in tickets:
                hours = ""
                if t.resolved_at and t.created_at:
                    hours = round((t.resolved_at - t.created_at).total_seconds() / 3600, 1)
                writer.writerow([t.reference_number, t.category, t.priority, t.status,
                                t.created_at.isoformat() if t.created_at else "",
                                t.resolved_at.isoformat() if t.resolved_at else "", hours])

        elif section == "agents":
            managers = User.query.filter(User.role.in_(["manager"])).all()
            writer.writerow(["Agent", "Resolved", "Pending", "Escalated", "Avg Hours", "Rating"])
            for mgr in managers:
                assigned = Ticket.query.filter(Ticket.assigned_to == mgr.id, Ticket.created_at >= start_date)
                resolved = assigned.filter(Ticket.status == "resolved").count()
                pending = assigned.filter(Ticket.status.in_(["pending", "in_progress"])).count()
                escalated = assigned.filter(Ticket.status == "escalated").count()
                writer.writerow([mgr.name, resolved, pending, escalated, 0, 0])

        elif section == "csat":
            feedbacks = Feedback.query.filter(Feedback.created_at >= start_date).all()
            writer.writerow(["User", "Rating", "Comment", "Date"])
            for f in feedbacks:
                writer.writerow([f.user.name if f.user else "", f.rating, f.comment,
                                f.created_at.isoformat() if f.created_at else ""])

        elif section == "sla":
            tickets = Ticket.query.filter(
                Ticket.created_at >= start_date, Ticket.status == "resolved",
                Ticket.resolved_at.isnot(None)
            ).all()
            sla_targets = get_sla_targets()
            writer.writerow(["Reference", "Priority", "Target Hours", "Actual Hours", "Status"])
            for t in tickets:
                hours = round((t.resolved_at - t.created_at).total_seconds() / 3600, 1)
                target = sla_targets.get(t.priority, 48)
                status = "Within SLA" if hours <= target else "Breached"
                writer.writerow([t.reference_number, t.priority, target, hours, status])

        response = Response(output.getvalue(), mimetype="text/csv")
        response.headers["Content-Disposition"] = f"attachment; filename=report_{section}_{range_param}.csv"
        return response

    return jsonify({"error": "PDF export is handled client-side"}), 400


# ═══════════════════════════════════════════════════════════════════════════════
# INIT DB + SEED ADMIN
# ═══════════════════════════════════════════════════════════════════════════════

with app.app_context():
    db.create_all()

    # Seed default admin if none exists
    if not User.query.filter_by(role="admin").first():
        admin = User(name="Admin", email="didardeep.12@gmail.com", role="admin", employee_id="ADM00001")
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()
        print(">>> Default admin created: didardeep.12@gmail.com / admin123")

    # Backfill employee_ids for existing non-customer users
    users_without_emp_id = User.query.filter(
        User.role != "customer",
        User.employee_id.is_(None)
    ).all()
    for u in users_without_emp_id:
        u.employee_id = generate_employee_id(u.role)
    if users_without_emp_id:
        db.session.commit()
        print(f">>> Backfilled employee_ids for {len(users_without_emp_id)} users")

    # Seed SLA defaults if not present
    for key, info in SLA_DEFAULTS.items():
        if not SystemSetting.query.filter_by(key=key).first():
            setting = SystemSetting(key=key, value=info["value"], category="sla", description=info["description"])
            db.session.add(setting)
    db.session.commit()


if __name__ == "__main__":
    app.run(debug=True, port=5500)
